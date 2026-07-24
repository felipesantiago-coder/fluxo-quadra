import sys, os
sys.path.insert(0, "/home/z/my-project/skills/xlsx/templates")
from base import *
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()
ws = wb.active
ws.title = "Modelo de Unidades"

# ─── Colunas esperadas pelo sistema ───────────────────────────────────────
# Ordem e nomes amigáveis em PT-BR (os nomes alternativos também funcionam)
headers = [
    ("Unidade", True, "Obrigatória. Identificador da unidade (ex: 101, A-201, U-03). Aceita: unidade, apto, nº unidade, apartamento"),
    ("Andar", False, "Numérico. Pavimento/andar da unidade. Aceita: andar, pavimento, floor"),
    ("Área Privativa", False, "Numérico, em m². Ex: 85.5. Aceita: área, área privativa, m², metragem"),
    ("Quartos", False, "Numérico inteiro. Ex: 3. Aceita: quartos, dormitórios, suítes"),
    ("Vagas", False, "Numérico inteiro. Ex: 2. Aceita: vagas, garagem, vaga"),
    ("Valor de Venda", False, "Numérico (R$). Use formato brasileiro: 1.234.567,89 ou 1234567.89. Aceita: valor, valor de venda, preço, preço de venda"),
    ("Status", False, "Texto. Valores aceitos: disponível, reservada, vendida. Padrão (vazio): disponível"),
    ("Posição Solar", False, "Texto. Ex: Norte, Leste, Sul, Oeste, NL. Aceita: posição solar, posição, face, sol"),
    ("Tipologia", False, "Texto. Ex: 2 quartos, Studio, Cobertura 4Q. Aceita: tipologia, tipo, tipo unidade, planta"),
    ("Bloco", False, "Texto. Ex: A, B, Torre 1. Aceita: bloco, torre"),
    ("Cobertura", False, "Sim/Não. Indica se é cobertura (penthouse). Aceita: cobertura"),
    ("Garden", False, "Sim/Não. Indica se é garden/terreo. Aceita: garden"),
]

last_col = len(headers) + 1  # +1 because we start at B=2

# ─── Setup da planilha ────────────────────────────────────────────────────
setup_sheet(ws, title="Modelo de Upload de Unidades", last_col=last_col)

# ─── Row 3: Subtítulo com instruções ───────────────────────────────────────────
ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=last_col)
subtitle_cell = ws["B3"]
subtitle_cell.value = "Preencha as colunas desejadas. Apenas \"Unidade\" é obrigatória. Unidades já existentes serão atualizadas (UPSERT)."
subtitle_cell.font = Font(name=FONT_NAME, size=10, color=NEUTRAL_600)
subtitle_cell.alignment = Alignment(horizontal='left', vertical='center')
ws.row_dimensions[3].height = 20

# ─── Row 4: Cabeçalhos ─────────────────────────────────────────────────────
for col_idx, (header_name, required, _) in enumerate(headers, start=2):
    cell = ws.cell(row=4, column=col_idx, value=header_name)
    if required:
        cell.value = f"{header_name} *"

style_header_row(ws, row_num=4, col_start=2, col_end=last_col)
ws.row_dimensions[4].height = 30

# ─── Row 5: Descrição / dicas (fonte caption, cinza) ────────────────────────
for col_idx, (_, _, tip) in enumerate(headers, start=2):
    cell = ws.cell(row=5, column=col_idx, value=tip)
    cell.font = font_caption()
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell.fill = PatternFill('solid', fgColor=PRIMARY_LIGHT)
ws.row_dimensions[5].height = 55

# ─── Rows 6-7: Exemplos preenchidos ────────────────────────────────────────
example_data = [
    ["101",  1,   65.2,  2, 1, 485000.00, "disponível", "Norte", "2 Quartos", "A", "Não", "Não"],
    ["201",  2,   82.7,  3, 2, 720000.00, "reservada",   "Leste", "3 Quartos", "A", "Não", "Não"],
    ["301",  3,  110.5,  4, 3, 1250000.00, "vendida",    "Sul",   "4 Quartos", "B", "Sim",  "Não"],
    ["G-01",  0,   95.0,  3, 2, 890000.00, "disponível", "Oeste", "3 Quartos", "A", "Não", "Sim"],
]

for i, row_data in enumerate(example_data):
    row_num = 6 + i
    for col_idx, value in enumerate(row_data, start=2):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
    style_data_row(ws, row_num=row_num, col_start=2, col_end=last_col, row_index=i)
    ws.row_dimensions[row_num].height = 22

# Formato brasileiro para Valor de Venda (coluna G = index 7, col_idx 7)
for row_num in range(6, 10):
    cell = ws.cell(row=row_num, column=7)  # Valor de Venda
    cell.number_format = '#,##0.00'
    cell.alignment = align_number()

# Formato numérico para Área (coluna D = col_idx 4)
for row_num in range(6, 10):
    cell = ws.cell(row=row_num, column=4)  # Área Privativa
    cell.number_format = '0.0'
    cell.alignment = align_number()

# Data validation para Status (coluna H = col_idx 8)
dv_status = DataValidation(
    type="list",
    formula1='"disponível,reservada,vendida"',
    allow_blank=True
)
dv_status.error = 'Valores aceitos: disponível, reservada, vendida'
dv_status.errorTitle = 'Status inválido'
ws.add_data_validation(dv_status)
dv_status.add(f'H6:H1000')

# Data validation para Cobertura e Garden (colunas L e M = col_idx 12 e 13)
dv_bool = DataValidation(
    type="list",
    formula1='"Sim,Não"',
    allow_blank=True
)
dv_bool.error = 'Use Sim ou Não'
dv_bool.errorTitle = 'Valor inválido'
ws.add_data_validation(dv_bool)
dv_bool.add(f'L6:L1000')
dv_bool.add(f'M6:M1000')

# ─── Row 11: Nota de rodapé ────────────────────────────────────────────────
footnote_row = 6 + len(example_data) + 1  # row 11
ws.merge_cells(start_row=footnote_row, start_column=2, end_row=footnote_row, end_column=last_col)
footnote = ws.cell(row=footnote_row, column=2)
footnote.value = "* Coluna obrigatória. As demais são opcionais. Exclua as linhas de exemplo antes de usar."
footnote.font = font_caption()
footnote.alignment = Alignment(horizontal='left', vertical='center')

# ─── Larguras de colunas ───────────────────────────────────────────────────
col_widths = {
    2: 14,   # Unidade
    3: 10,   # Andar
    4: 16,   # Área Privativa
    5: 10,   # Quartos
    6: 10,   # Vagas
    7: 20,   # Valor de Venda
    8: 16,   # Status
    9: 18,   # Posição Solar
    10: 20,  # Tipologia
    11: 12,  # Bloco
    12: 12,  # Cobertura
    13: 12,  # Garden
}
for col, width in col_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = width

# Coluna A como margem
ws.column_dimensions['A'].width = 3

# Grid lines ocultas
ws.sheet_view.showGridLines = False

# Freeze panes: cabeçalhos fixos
ws.freeze_panes = 'B6'

# ─── Salvar ────────────────────────────────────────────────────────────────
output_path = "/home/z/my-project/download/template_upload_unidades.xlsx"
os.makedirs(os.path.dirname(output_path), exist_ok=True)
wb.properties.creator = "Z.ai"
wb.save(output_path)
print(f"Template salvo em: {output_path}")